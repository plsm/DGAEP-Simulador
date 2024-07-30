import argparse
import csv

import openpyxl


class Args:
    def __init__(self):
        parser = argparse.ArgumentParser (
            description='Processa o ficheiro EXCEL com os dados do boletim estatístico do emprego público e constroi '
                        'um ficheiro CSV com os postos de trabalho organizados por cargo, carreira ou grupo.'
        )
        parser.add_argument (
            'ficheiro_excel',
            type=str,
            help='Nome do ficheiro EXCEL.'
        )
        _args = parser.parse_args ()
        self.nome_ficheiro = _args.ficheiro_excel  # type: str


class Administracao:
    def __init__ (self, nome_excel, nome_csv):
        """
        Caracterização das diferentes administrações tal como estão no ficheiro Excel do BOEP
        :param nome_excel: nome da administração usado no Excel.
        :param nome_csv: Nome da administração a escrever no CSV.
        """
        self.nome_excel = nome_excel  # type: str
        self.nome_csv = nome_csv  # type: str


args = Args ()
folha_calculo = openpyxl.load_workbook (
    filename=args.nome_ficheiro
)

ADMNISTRACOES = [
    Administracao ('ADMINISTRAÇÃO CENTRAL ',            'Administração Central'),
    Administracao ('ADMINISTRAÇÃO REGIONAL DOS AÇORES', 'Administração Regional dos Açores'),
    Administracao ('ADMINISTRAÇÃO REGIONAL DA MADEIRA', 'Administração Regional da Madeira'),
    Administracao ('ADMINISTRAÇÃO LOCAL',               'Administração Local'),
    Administracao ('FUNDOS DE SEGURANÇA SOCIAL',        'Fundos de Segurança Social'),
]

CARGOS_IGNORAR = [
    'Dirigente superior:',
    'Dirigente intermédio:'
]

DATA_PARTIDA = 2011.5
NUMERO_FAIXAS_ETARIAS = 6

with open ('postos-trabalho-por-cargo-carreira-grupo.csv', 'w') as fd:
    writer = csv.writer (fd, quoting=csv.QUOTE_NONNUMERIC)
    writer.writerow ([
        'tempo',
        'administração',
        'cargo',
        'faixa_etária',
        'sexo',
        'postos_de_trabalho',
    ])
    indice_quadro = 1
    while True:
        nome_folha = f'Q.1.3.{indice_quadro}'
        if nome_folha not in folha_calculo.sheetnames:
            print (f'Folha {nome_folha} não existe na folha de cálculo {args.nome_ficheiro}.')
            break
        print (f'Processar os dados na folha {nome_folha}...')
        tempo = DATA_PARTIDA + (indice_quadro - 1) / 2
        a_folha = folha_calculo [nome_folha]
        numero_linha = 8
        for a_administracao in ADMNISTRACOES:
            # procurar a linha onde começam os dados da administração
            while a_folha.cell (row=numero_linha, column=2).value != a_administracao.nome_excel and \
                    a_folha.cell (numero_linha, column=3).value != a_administracao.nome_excel and \
                    numero_linha < a_folha.max_row:
                numero_linha += 1
            # saltar a linha com o total
            numero_linha += 1
            while a_folha.cell (row=numero_linha, column=4).value is not None:
                nome_cargo = a_folha.cell (row=numero_linha, column=4).value
                # ignorar linhas que são agregações de outros cargos
                if nome_cargo in CARGOS_IGNORAR:
                    numero_linha += 1
                    continue
                for indice_faixa_etaria in range (NUMERO_FAIXAS_ETARIAS):
                    for indice_sexo, nome_sexo in zip (range (2), ['H', 'M']):
                        numero_postos = a_folha.cell (
                            row=numero_linha,
                            column=5 + indice_faixa_etaria * 3 + indice_sexo
                        ).value
                        linha_escrever = [
                            tempo,
                            a_administracao.nome_csv,
                            nome_cargo,
                            indice_faixa_etaria,
                            nome_sexo,
                            numero_postos,
                        ]
                        writer.writerow (linha_escrever)
                numero_linha += 1
        indice_quadro += 1
