import csv
import sys

import openpyxl


class Administracao:
    def __init__ (self, nome, linha_inicial_ministerio_secretaria, linha_final_ministerio_secretaria):
        self.nome = nome
        self.linha_inicial_ministerio_secretaria = linha_inicial_ministerio_secretaria
        self.linha_final_ministerio_secretaria = linha_final_ministerio_secretaria


nome_ficheiro = sys.argv [1]
folha_calculo = openpyxl.load_workbook (
    filename=nome_ficheiro
)

ADMINISTRACOES = [
    Administracao ('Administração Central',             11, 34),
    Administracao ('Administração Regional dos Açores', 38, 51),
    Administracao ('Administração Regional da Madeira', 53, 67),
    Administracao ('Administração Local',               69, 74),
    Administracao ('Fundos de Segurança Social',        77, 79),
]
MINISTERIOS_SECRETARIAS_IGNORAR = [
    'Estado',
    'Serviços e Fundos Autónomos',
    'Estado e Serviços e Fundos Autónomos',
    'Órgãos do Governo Regional dos Açores',
    'Serviços e Fundos Autónomos da AR dos Açores',
    'Órgãos do Governo Regional da Madeira',
    'Serviços e Fundos Autónomos da AR da Madeira',
    'dos quais: Sector Empresarial Local - Entidades Reclassif. (ii)',
]
DATA_PARTIDA = 2011.5
NUMERO_FAIXAS_ETARIAS = 6

with open ('posto-trabalho-por-subsetor-ministerios-secretarias.csv', 'w') as fd:
    writer = csv.writer (
        fd,
        quoting=csv.QUOTE_NONNUMERIC
    )
    writer.writerow ([
        'tempo',
        'administração',
        'ministério_secretaria',
        'faixa_etária',
        'sexo',
        'postos de trabalho'
    ])
    indice_quadro = 1
    while True:
        nome_folha = f'Q.1.1.{indice_quadro}'
        if nome_folha not in folha_calculo.sheetnames:
            print (f'Folha {nome_folha} não existe na folha de cálculo {nome_ficheiro}.')
            break
        print (nome_folha)
        tempo = DATA_PARTIDA + (indice_quadro - 1) / 2
        a_folha = folha_calculo [nome_folha]
        for administracao in ADMINISTRACOES:
            for linha_ministerio_secretaria in range (
                    administracao.linha_inicial_ministerio_secretaria,
                    administracao.linha_final_ministerio_secretaria + 1):
                nome_ministerio_secretaria = a_folha.cell (row=linha_ministerio_secretaria, column=2).value
                if nome_ministerio_secretaria in MINISTERIOS_SECRETARIAS_IGNORAR:
                    continue
                for indice_faixa_etaria in range (NUMERO_FAIXAS_ETARIAS):
                    for indice_sexo, nome_sexo in zip (range (2), ['H', 'M']):
                        celula = a_folha.cell(
                            row=linha_ministerio_secretaria,
                            column=3 + indice_faixa_etaria * 3 + indice_sexo
                        )
                        linha = [
                            tempo,
                            administracao.nome,
                            nome_ministerio_secretaria,
                            indice_faixa_etaria,
                            nome_sexo,
                            celula.value,
                        ]
                        writer.writerow (linha)
            print (f'   {administracao.nome} @ {administracao.linha_inicial_ministerio_secretaria} ->'
                   f' {administracao.linha_final_ministerio_secretaria}')
        indice_quadro += 1
